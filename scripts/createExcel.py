#!/usr/bin/env python3
"""Parse HTTP KV server load test results and append metrics to an Excel sheet."""
from __future__ import annotations

import argparse
import os
import re
from typing import Iterable, List, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ModuleNotFoundError as exc:  # pragma: no cover - guides user to install dependency
    raise SystemExit(
        "Missing dependency 'openpyxl'. Install it with 'pip install openpyxl' and retry."
    ) from exc


RESULT_PATTERN = re.compile(r"^\s*Clients:\s*(?P<clients>\d+)")
THROUGHPUT_PATTERN = re.compile(r"^\s*Throughput \(req/s\):\s*(?P<throughput>[0-9]*\.?[0-9]+)")
LATENCY_PATTERN = re.compile(r"^\s*Avg latency \(ms\):\s*(?P<latency>[0-9]*\.?[0-9]+)")


def parse_results(lines: Iterable[str]) -> List[Tuple[int, float, float]]:
    """Extract (clients, throughput, latency_ms) tuples from the log lines."""
    records: List[Tuple[int, float, float]] = []
    clients: int | None = None
    throughput: float | None = None
    latency: float | None = None

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue

        match_clients = RESULT_PATTERN.match(line)
        if match_clients:
            clients = int(match_clients.group("clients"))
            continue

        match_throughput = THROUGHPUT_PATTERN.match(line)
        if match_throughput:
            throughput = float(match_throughput.group("throughput"))
            continue

        match_latency = LATENCY_PATTERN.match(line)
        if match_latency:
            latency = float(match_latency.group("latency"))

        if clients is not None and throughput is not None and latency is not None:
            records.append((clients, throughput, latency))
            clients = throughput = latency = None

    return records


def ensure_sheet(workbook_path: str, sheet_name: str) -> Tuple[Workbook, Worksheet, bool]:
    """Load or create a workbook and return the target worksheet with header readiness."""
    if os.path.exists(workbook_path):
        wb = load_workbook(workbook_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
        needs_header = True
        ws.delete_rows(1)
    else:
        needs_header = False
    return wb, ws, needs_header


def append_to_sheet(ws: Worksheet, records: List[Tuple[int, float, float]], add_header: bool) -> None:
    """Append the parsed records to the worksheet."""
    if add_header:
        ws.append(["Clients", "Throughput (req/s)", "Avg latency (ms)"])

    for clients, throughput, latency in records:
        ws.append([clients, throughput, latency])


def main() -> None:
    parser = argparse.ArgumentParser(description="Append parsed metrics to an Excel workbook.")
    parser.add_argument("logfile", help="Path to the results log file to parse.")
    parser.add_argument(
        "--workbook",
        default=None,
        help="Excel workbook path to append to (default: results/load_test_results.xlsx)",
    )
    parser.add_argument(
        "--sheet",
        default="Results",
        help="Worksheet name to append to (default: Results)",
    )
    args = parser.parse_args()

    if args.workbook is None:
        if os.path.basename(os.getcwd()) == "scripts" and os.path.isdir("../results"):
            args.workbook = "../results/load_test_results.xlsx"
        else:
            args.workbook = "results/load_test_results.xlsx"

    # Ensure results directory exists if workbook is in it
    wb_dir = os.path.dirname(args.workbook)
    if wb_dir:
        os.makedirs(wb_dir, exist_ok=True)

    # Check if logfile exists, if not check in results/
    logfile = args.logfile
    if not os.path.exists(logfile) and not os.path.isabs(logfile):
        candidate = os.path.join("results", logfile)
        if os.path.exists(candidate):
            logfile = candidate

    with open(logfile, "r", encoding="utf-8") as f:
        records = parse_results(f)

    if not records:
        raise SystemExit("No result blocks were found in the provided file.")

    wb, ws, needs_header = ensure_sheet(args.workbook, args.sheet)
    append_to_sheet(ws, records, needs_header)
    wb.save(args.workbook)


if __name__ == "__main__":
    main()
